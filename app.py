# --- START OF FILE app.py ---

from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from pyXSteam.XSteam import XSteam
import io
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import os
import sys
import math

# Steam Tablosu birimlerini ayarla
steamTable = XSteam(XSteam.UNIT_SYSTEM_MKS)

app = Flask(__name__)
CORS(app)

# =============================================================================
# YARDIMCI FONKSİYONLAR
# =============================================================================

def kpa_to_bar(p_kpa):
    try: return float(p_kpa) / 100.0
    except: raise ValueError(f"Invalid pressure value for kPa to bar conversion: {p_kpa}")

def bar_to_kpa(p_bar):
    try: return float(p_bar) * 100.0
    except: raise ValueError(f"Invalid pressure value for bar to kPa conversion: {p_bar}")

def get_phase_from_quality(x):
    epsilon = 1e-6
    if x is None or (isinstance(x, float) and math.isnan(x)) or x < 0 - epsilon or x > 1 + epsilon: return "Invalid Quality", None
    if abs(x - 0) < epsilon: return f"Saturated Liquid (x={x:.4e})", 0.0
    if abs(x - 1) < epsilon: return f"Saturated Vapor (x={x:.4e})", 1.0
    if 0 + epsilon <= x <= 1 - epsilon: return f"Two Phase Mixture (x={x:.4f})", x
    return "Invalid Quality", None

def get_phase_from_ph(p_bar, h_val):
    epsilon = 1e-6
    if math.isnan(p_bar) or math.isnan(h_val): return "Invalid Input", None
    try:
        hf = steamTable.hL_p(p_bar)
        hg = steamTable.hV_p(p_bar)
        if h_val < hf - epsilon: return "Subcooled Liquid", None
        if abs(h_val - hf) < epsilon: return f"Saturated Liquid (x={0.0:.4e})", 0.0
        if h_val > hg + epsilon: return "Superheated Vapor", None
        if abs(h_val - hg) < epsilon: return f"Saturated Vapor (x={1.0:.4e})", 1.0
        quality = (h_val - hf) / (hg - hf)
        return get_phase_from_quality(quality)
    except: return "Supercritical or out of range", None

def sanitize_for_json(props):
    sanitized = {}
    for key, value in props.items():
        if value is None or (isinstance(value, float) and math.isnan(value)): sanitized[key] = None
        else: sanitized[key] = value
    return sanitized

def safe_calculate(calculation_func, *args):
    original_stdout, original_stderr = sys.stdout, sys.stderr
    sys.stdout, sys.stderr = open(os.devnull, 'w'), open(os.devnull, 'w')
    try:
        result = calculation_func(*args)
        return float('nan') if result is None else result
    except Exception: return float('nan')
    finally:
        sys.stdout.close(); sys.stderr.close()
        sys.stdout, sys.stderr = original_stdout, original_stderr

def get_saturation_properties(p_kpa=None, t_c=None):
    """Calculate saturation properties at given pressure or temperature"""
    sat_props = {}
    try:
        if p_kpa is not None:
            p_bar = kpa_to_bar(p_kpa)
            sat_props['h_f'] = safe_calculate(steamTable.hL_p, p_bar)
            sat_props['h_g'] = safe_calculate(steamTable.hV_p, p_bar)
            sat_props['h_fg'] = sat_props['h_g'] - sat_props['h_f'] if not (math.isnan(sat_props['h_f']) or math.isnan(sat_props['h_g'])) else float('nan')
            sat_props['s_f'] = safe_calculate(steamTable.sL_p, p_bar)
            sat_props['s_g'] = safe_calculate(steamTable.sV_p, p_bar)
            sat_props['s_fg'] = sat_props['s_g'] - sat_props['s_f'] if not (math.isnan(sat_props['s_f']) or math.isnan(sat_props['s_g'])) else float('nan')
            sat_props['v_f'] = safe_calculate(steamTable.vL_p, p_bar)
            sat_props['v_g'] = safe_calculate(steamTable.vV_p, p_bar)
            sat_props['v_fg'] = sat_props['v_g'] - sat_props['v_f'] if not (math.isnan(sat_props['v_f']) or math.isnan(sat_props['v_g'])) else float('nan')
            sat_props['T_sat'] = safe_calculate(steamTable.tsat_p, p_bar)
        elif t_c is not None:
            sat_props['h_f'] = safe_calculate(steamTable.hL_t, t_c)
            sat_props['h_g'] = safe_calculate(steamTable.hV_t, t_c)
            sat_props['h_fg'] = sat_props['h_g'] - sat_props['h_f'] if not (math.isnan(sat_props['h_f']) or math.isnan(sat_props['h_g'])) else float('nan')
            sat_props['s_f'] = safe_calculate(steamTable.sL_t, t_c)
            sat_props['s_g'] = safe_calculate(steamTable.sV_t, t_c)
            sat_props['s_fg'] = sat_props['s_g'] - sat_props['s_f'] if not (math.isnan(sat_props['s_f']) or math.isnan(sat_props['s_g'])) else float('nan')
            sat_props['v_f'] = safe_calculate(steamTable.vL_t, t_c)
            sat_props['v_g'] = safe_calculate(steamTable.vV_t, t_c)
            sat_props['v_fg'] = sat_props['v_g'] - sat_props['v_f'] if not (math.isnan(sat_props['v_f']) or math.isnan(sat_props['v_g'])) else float('nan')
            sat_props['P_sat'] = safe_calculate(steamTable.psat_t, t_c)
            if not math.isnan(sat_props['P_sat']):
                sat_props['P_sat'] = bar_to_kpa(sat_props['P_sat'])
    except:
        pass
    return sat_props

def resolve_string_value(prop_type, prop_value_str, other_prop, other_value):
    prop_value_str = prop_value_str.lower()
    if other_prop == 'p': # Gelen verinin küçük harf olacağını varsayıyoruz
        p_bar = kpa_to_bar(other_value)
        mapping = { 'h': {'hf': steamTable.hL_p, 'hg': steamTable.hV_p}, 's': {'sf': steamTable.sL_p, 'sg': steamTable.sV_p}, 'v': {'vf': steamTable.vL_p, 'vg': steamTable.vV_p} }
        if prop_type in mapping and prop_value_str in mapping[prop_type]: return safe_calculate(mapping[prop_type][prop_value_str], p_bar)
    elif other_prop == 't': # Gelen verinin küçük harf olacağını varsayıyoruz
        t_c = float(other_value)
        mapping = { 'h': {'hf': steamTable.hL_t, 'hg': steamTable.hV_t}, 's': {'sf': steamTable.sL_t, 'sg': steamTable.sV_t}, 'v': {'vf': steamTable.vL_t, 'vg': steamTable.vV_t} }
        if prop_type in mapping and prop_value_str in mapping[prop_type]: return safe_calculate(mapping[prop_type][prop_value_str], t_c)
    raise ValueError(f"Cannot resolve string value '{prop_value_str}' for property '{prop_type}' with given pair.")

# =============================================================================
# HAL ÖZELLİĞİ HESAPLAMA FONKSİYONLARI
# =============================================================================

def calc_from_px(p_kpa, x_val):
    p_bar, x = kpa_to_bar(p_kpa), float(x_val); phase, x = get_phase_from_quality(x); t = safe_calculate(steamTable.tsat_p, p_bar); h = safe_calculate(steamTable.h_px, p_bar, x); sL, sV = safe_calculate(steamTable.sL_p, p_bar), safe_calculate(steamTable.sV_p, p_bar); s = sL + x * (sV - sL); vL, vV = safe_calculate(steamTable.vL_p, p_bar), safe_calculate(steamTable.vV_p, p_bar); v = vL + x * (vV - vL); sat_props = get_saturation_properties(p_kpa=p_kpa); return {"T": t, "s": s, "P": float(p_kpa), "h": h, "v": v, "x": x, "phase": phase, "saturation": sat_props}
def calc_from_tx(t_c, x_val):
    t, x = float(t_c), float(x_val); phase, x = get_phase_from_quality(x); p_bar = safe_calculate(steamTable.psat_t, t); h = safe_calculate(steamTable.h_tx, t, x); vL, vV = safe_calculate(steamTable.vL_t, t), safe_calculate(steamTable.vV_t, t); v = vL + x * (vV - vL); sf, sg = safe_calculate(steamTable.sL_t, t), safe_calculate(steamTable.sV_t, t); s = sf + x * (sg - sf); sat_props = get_saturation_properties(t_c=t); return {"T": t, "s": s, "P": bar_to_kpa(p_bar), "h": h, "v": v, "x": x, "phase": phase, "saturation": sat_props}
def calc_from_ph(p_kpa, h_val):
    p_bar, h = kpa_to_bar(p_kpa), float(h_val); t = safe_calculate(steamTable.t_ph, p_bar, h); s = safe_calculate(steamTable.s_ph, p_bar, h); v = safe_calculate(steamTable.v_ph, p_bar, h); phase, x = get_phase_from_ph(p_bar, h); sat_props = get_saturation_properties(p_kpa=p_kpa); return {"T": t, "s": s, "P": float(p_kpa), "h": h, "v": v, "x": x, "phase": phase, "saturation": sat_props}
def calc_from_ps(p_kpa, s_val):
    p_bar, s = kpa_to_bar(p_kpa), float(s_val); t = safe_calculate(steamTable.t_ps, p_bar, s); h = safe_calculate(steamTable.h_ps, p_bar, s); v = safe_calculate(steamTable.v_ps, p_bar, s); phase, x = get_phase_from_ph(p_bar, h); sat_props = get_saturation_properties(p_kpa=p_kpa); return {"T": t, "s": s, "P": float(p_kpa), "h": h, "v": v, "x": x, "phase": phase, "saturation": sat_props}
def calc_from_pt(p_kpa, t_c):
    p_bar, t = kpa_to_bar(p_kpa), float(t_c); tsat = safe_calculate(steamTable.tsat_p, p_bar);
    if not math.isnan(tsat) and abs(t - tsat) < 1e-4: raise ValueError("Input properties are dependent (T is saturation temperature for P).")
    h = safe_calculate(steamTable.h_pt, p_bar, t); s = safe_calculate(steamTable.s_pt, p_bar, t); v = safe_calculate(steamTable.v_pt, p_bar, t); phase, x = get_phase_from_ph(p_bar, h); sat_props = get_saturation_properties(p_kpa=p_kpa); return {"T": t, "s": s, "P": float(p_kpa), "h": h, "v": v, "x": x, "phase": phase, "saturation": sat_props}

# =============================================================================
# API ENDPOINTLERİ
# =============================================================================

# app.py dosyanızdaki mevcut /calculate_from_pairs fonksiyonunu bununla değiştirin.

@app.route('/calculate_from_pairs', methods=['POST'])
def calculate_from_pairs():
    data = request.get_json()
    if not data or 'prop1' not in data or 'prop2' not in data:
        return jsonify({"error": "Invalid data format"}), 400
    
    try:
        name = data.get("name", "Point")
        prop1 = data.get("prop1")
        prop2 = data.get("prop2")

        # --- YENİ VE BASİT MANTIK ---
        # 1. Gelen veriyi bir sözlüğe (dictionary) atayalım. Tipleri küçük harfe çevirelim.
        props_map = {
            prop1['type'].lower(): prop1['value'],
            prop2['type'].lower(): prop2['value']
        }

        # 2. String değerleri (hf, hg vb.) sayısal değerlere dönüştürelim
        for p_type, p_val in list(props_map.items()): # list() kopyası üzerinde döngü
            if isinstance(p_val, str):
                other_types = list(props_map.keys())
                other_types.remove(p_type)
                if not other_types:
                    raise ValueError("Cannot resolve string value without a second property.")
                other_type = other_types[0]
                # resolve_string_value'ya küçük harf tiplerle gidelim
                props_map[p_type] = resolve_string_value(p_type, p_val, other_type, props_map[other_type])

        # 3. Hangi fonksiyonu çağıracağımızı basitçe kontrol edelim ve çağıralım
        properties = {}
        if 'p' in props_map and 'h' in props_map:
            properties = calc_from_ph(props_map['p'], props_map['h'])
        elif 'p' in props_map and 's' in props_map:
            properties = calc_from_ps(props_map['p'], props_map['s'])
        elif 'p' in props_map and 't' in props_map:
            properties = calc_from_pt(props_map['p'], props_map['t'])
        elif 'p' in props_map and 'x' in props_map:
            properties = calc_from_px(props_map['p'], props_map['x'])
        elif 't' in props_map and 'x' in props_map:
            properties = calc_from_tx(props_map['t'], props_map['x'])
        else:
            # Eğer hiçbir koşul eşleşmezse, desteklenmeyen çift hatası ver
            keys = list(props_map.keys())
            return jsonify({"error": f"Unsupported property pair: {keys[0]}, {keys[1]}"}), 400
        
        properties["name"] = name
        
        sanitized_properties = sanitize_for_json(properties)
        if sanitized_properties.get('T') is None or sanitized_properties.get('s') is None:
             raise ValueError("Could not calculate T/s. Inputs may be out of valid range.")
        
        return jsonify({"point": sanitized_properties}), 200
        
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Calculation failed: {str(e)}"}), 500

# app.py dosyanızdaki mevcut export_xlsx fonksiyonunu bununla değiştirin.

@app.route('/export_xlsx', methods=['POST'])
def export_xlsx():
    try:
        points = request.get_json().get('points', [])
        if not points: 
            return jsonify({"error": "No data to export"}), 400
        
        wb = Workbook()
        default_text = "Not Available"
        
        # ===============================================
        # SHEET 1: THERMODYNAMIC PROPERTIES (Ana Tablo)
        # ===============================================
        ws1 = wb.active
        ws1.title = "Thermodynamic Properties"
        
        headers = ['Point Name', 'Phase', 'Pressure (kPa)', 'Temperature (°C)', 'Enthalpy (kJ/kg)', 'Entropy (kJ/kg·K)', 'Spec. Vol. (m³/kg)', 'Quality']
        ws1.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Header satır yüksekliği
        ws1.row_dimensions[1].height = 25

        for point in points:
            quality = point.get('x')
            if quality is None:
                quality_str = default_text
            else:
                quality_str = f"{quality:.4f}"

            row = [
                point.get('name', default_text),
                point.get('phase', default_text),
                point.get('P', default_text),
                point.get('T', default_text),
                point.get('h', default_text),
                point.get('s', default_text),
                point.get('v', default_text),
                quality_str
            ]
            ws1.append(row)
        
        # Sütun genişliklerini ve hizalamayı ayarla
        for col_idx, column_cells in enumerate(ws1.columns, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            
            for cell in column_cells:
                # Tüm hücreleri ortala
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except:
                    pass
            
            # Genişliği ayarla (minimum 12, maksimum 25)
            adjusted_width = min(max(max_length + 2, 12), 25)
            ws1.column_dimensions[column].width = adjusted_width
        
        # ===============================================
        # SHEET 2: SATURATION PROPERTIES
        # ===============================================
        points_with_sat = [p for p in points if p.get('saturation') and isinstance(p.get('saturation'), dict) and len(p.get('saturation')) > 0]
        
        if points_with_sat:
            ws2 = wb.create_sheet(title="Saturation Properties")
            
            # Başlık
            ws2.cell(row=1, column=1, value="SATURATION PROPERTIES")
            title_cell = ws2.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
            
            # Tablo başlıkları
            sat_headers = ['Point', 'hf (kJ/kg)', 'hfg (kJ/kg)', 'hg (kJ/kg)', 
                          'sf (kJ/kg·K)', 'sfg (kJ/kg·K)', 'sg (kJ/kg·K)',
                          'vf (m³/kg)', 'vfg (m³/kg)', 'vg (m³/kg)', 'Tsat/Psat']
            
            for col_idx, header in enumerate(sat_headers, 1):
                cell = ws2.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Header satır yükseklikleri
            ws2.row_dimensions[1].height = 25  # Başlık satırı
            ws2.row_dimensions[3].height = 30  # Tablo başlıkları
            
            current_row = 4
            for point in points_with_sat:
                sat = point.get('saturation', {})
                
                tsat_psat = ""
                if 'T_sat' in sat and not math.isnan(sat['T_sat']):
                    tsat_psat = f"Tsat={sat['T_sat']:.2f}°C"
                elif 'P_sat' in sat and not math.isnan(sat['P_sat']):
                    tsat_psat = f"Psat={sat['P_sat']:.2f}kPa"
                
                sat_row = [
                    point.get('name', default_text),
                    sat.get('h_f', default_text) if not math.isnan(sat.get('h_f', float('nan'))) else default_text,
                    sat.get('h_fg', default_text) if not math.isnan(sat.get('h_fg', float('nan'))) else default_text,
                    sat.get('h_g', default_text) if not math.isnan(sat.get('h_g', float('nan'))) else default_text,
                    sat.get('s_f', default_text) if not math.isnan(sat.get('s_f', float('nan'))) else default_text,
                    sat.get('s_fg', default_text) if not math.isnan(sat.get('s_fg', float('nan'))) else default_text,
                    sat.get('s_g', default_text) if not math.isnan(sat.get('s_g', float('nan'))) else default_text,
                    sat.get('v_f', default_text) if not math.isnan(sat.get('v_f', float('nan'))) else default_text,
                    sat.get('v_fg', default_text) if not math.isnan(sat.get('v_fg', float('nan'))) else default_text,
                    sat.get('v_g', default_text) if not math.isnan(sat.get('v_g', float('nan'))) else default_text,
                    tsat_psat or default_text
                ]
                
                for col_idx, value in enumerate(sat_row, 1):
                    cell = ws2.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
            
            # Sütun genişliklerini otomatik ayarla
            for col_idx, column_cells in enumerate(ws2.columns, 1):
                max_length = 0
                column = get_column_letter(col_idx)
                
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass
                
                # Genişliği ayarla (minimum 14, maksimum 20)
                adjusted_width = min(max(max_length + 2, 14), 20)
                ws2.column_dimensions[column].width = adjusted_width
        
        # ===============================================
        # SHEET 3: DETAILED COMPARISON & ANALYSIS
        # ===============================================
        if points_with_sat:
            ws3 = wb.create_sheet(title="Detailed Analysis")
            
            # Başlık
            ws3.cell(row=1, column=1, value="DETAILED COMPARISON & ANALYSIS")
            title_cell = ws3.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
            
            # Başlık satır yüksekliği
            ws3.row_dimensions[1].height = 25
            
            # Tablo başlıkları
            analysis_headers = [
                'Point', 'Phase', 'Quality (x)',
                'h (Actual)', 'hf', 'hg', 'Δh from hf', 'Δh from hg', '% Between hf-hg',
                's (Actual)', 'sf', 'sg', 'Δs from sf', 'Δs from sg', '% Between sf-sg',
                'Region'
            ]
            
            for col_idx, header in enumerate(analysis_headers, 1):
                cell = ws3.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            current_row = 4
            for point in points_with_sat:
                sat = point.get('saturation', {})
                
                # Actual değerler
                h_actual = point.get('h')
                s_actual = point.get('s')
                v_actual = point.get('v')
                x_actual = point.get('x')
                
                # Saturation değerleri
                hf = sat.get('h_f')
                hg = sat.get('h_g')
                sf = sat.get('s_f')
                sg = sat.get('s_g')
                
                # Hesaplamalar
                delta_h_from_hf = h_actual - hf if (h_actual is not None and hf is not None and not math.isnan(h_actual) and not math.isnan(hf)) else None
                delta_h_from_hg = h_actual - hg if (h_actual is not None and hg is not None and not math.isnan(h_actual) and not math.isnan(hg)) else None
                
                percent_h = None
                if delta_h_from_hf is not None and hf is not None and hg is not None and not math.isnan(hf) and not math.isnan(hg):
                    percent_h = (delta_h_from_hf / (hg - hf)) * 100 if (hg - hf) != 0 else None
                
                delta_s_from_sf = s_actual - sf if (s_actual is not None and sf is not None and not math.isnan(s_actual) and not math.isnan(sf)) else None
                delta_s_from_sg = s_actual - sg if (s_actual is not None and sg is not None and not math.isnan(s_actual) and not math.isnan(sg)) else None
                
                percent_s = None
                if delta_s_from_sf is not None and sf is not None and sg is not None and not math.isnan(sf) and not math.isnan(sg):
                    percent_s = (delta_s_from_sf / (sg - sf)) * 100 if (sg - sf) != 0 else None
                
                # Region belirleme
                region = "Unknown"
                phase = point.get('phase', '')
                if 'Subcooled' in phase:
                    region = "Subcooled Liquid"
                elif 'Superheated' in phase:
                    region = "Superheated Vapor"
                elif 'Saturated Liquid' in phase:
                    region = "Saturated Liquid (x≈0)"
                elif 'Saturated Vapor' in phase:
                    region = "Saturated Vapor (x≈1)"
                elif 'Two Phase' in phase:
                    region = "Two-Phase Region"
                
                # Format helper
                def fmt_val(val, precision=4):
                    if val is None or (isinstance(val, float) and math.isnan(val)):
                        return default_text
                    return round(val, precision)
                
                analysis_row = [
                    point.get('name', default_text),
                    point.get('phase', default_text),
                    fmt_val(x_actual, 4),
                    fmt_val(h_actual, 4),
                    fmt_val(hf, 4),
                    fmt_val(hg, 4),
                    fmt_val(delta_h_from_hf, 4),
                    fmt_val(delta_h_from_hg, 4),
                    fmt_val(percent_h, 2),
                    fmt_val(s_actual, 4),
                    fmt_val(sf, 4),
                    fmt_val(sg, 4),
                    fmt_val(delta_s_from_sf, 4),
                    fmt_val(delta_s_from_sg, 4),
                    fmt_val(percent_s, 2),
                    region
                ]
                
                for col_idx, value in enumerate(analysis_row, 1):
                    cell = ws3.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Conditional formatting için renk
                    if col_idx == 16:  # Region sütunu
                        if "Subcooled" in str(value):
                            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                        elif "Superheated" in str(value):
                            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
                        elif "Two-Phase" in str(value):
                            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        elif "Saturated" in str(value):
                            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                
                current_row += 1
            
            # Sütun genişliklerini otomatik ve akıllı ayarla
            column_widths = {
                1: 10,  # Point
                2: 25,  # Phase
                3: 12,  # Quality
                4: 13,  # h (Actual)
                5: 13,  # hf
                6: 13,  # hg
                7: 13,  # Δh from hf
                8: 13,  # Δh from hg
                9: 15,  # % Between hf-hg
                10: 13, # s (Actual)
                11: 13, # sf
                12: 13, # sg
                13: 13, # Δs from sf
                14: 13, # Δs from sg
                15: 15, # % Between sf-sg
                16: 20  # Region
            }
            
            for col_idx, width in column_widths.items():
                ws3.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Tüm header hücrelerini wrap text yap
            for col_idx in range(1, 17):
                header_cell = ws3.cell(row=3, column=col_idx)
                header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Açıklama notları ekle
            ws3.cell(row=current_row + 2, column=1, value="NOTES:")
            notes_cell = ws3.cell(row=current_row + 2, column=1)
            notes_cell.font = Font(bold=True, size=11)
            notes_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws3.cell(row=current_row + 3, column=1, value="• '% Between hf-hg' shows position in two-phase region: (h - hf) / (hg - hf) × 100")
            ws3.cell(row=current_row + 4, column=1, value="• '% Between sf-sg' shows entropy position in two-phase region: (s - sf) / (sg - sf) × 100")
            ws3.cell(row=current_row + 5, column=1, value="• These percentages indicate how far the point is from saturated liquid (0%) to saturated vapor (100%)")
            
            for row_idx in range(current_row + 3, current_row + 6):
                cell = ws3.cell(row=row_idx, column=1)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font = Font(size=9, italic=True)
            
            # Merge notes cells
            ws3.merge_cells(start_row=current_row + 3, start_column=1, end_row=current_row + 3, end_column=16)
            ws3.merge_cells(start_row=current_row + 4, start_column=1, end_row=current_row + 4, end_column=16)
            ws3.merge_cells(start_row=current_row + 5, start_column=1, end_row=current_row + 5, end_column=16)
            
            # LEGEND
            ws3.cell(row=current_row + 7, column=1, value="LEGEND:")
            legend_cell = ws3.cell(row=current_row + 7, column=1)
            legend_cell.font = Font(bold=True, size=11)
            legend_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            legend_data = [
                ("Subcooled Liquid", "Green", "C6E0B4"),
                ("Superheated Vapor", "Orange", "F8CBAD"),
                ("Two-Phase Region", "Yellow", "FFE699"),
                ("Saturated States", "Blue", "BDD7EE")
            ]
            
            legend_row = current_row + 8
            for region_name, color_name, color_hex in legend_data:
                name_cell = ws3.cell(row=legend_row, column=1, value=region_name)
                name_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                color_cell = ws3.cell(row=legend_row, column=2, value=color_name)
                color_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                color_cell.alignment = Alignment(horizontal="center", vertical="center")
                color_cell.font = Font(bold=True)
                
                legend_row += 1
            
            # Legend sütun genişlikleri
            ws3.column_dimensions['A'].width = max(ws3.column_dimensions['A'].width, 20)
            ws3.column_dimensions['B'].width = max(ws3.column_dimensions['B'].width, 12)
            
            # Header satır yüksekliğini artır (daha iyi görünüm için)
            ws3.row_dimensions[3].height = 30
        
        mem_file = io.BytesIO()
        wb.save(mem_file)
        mem_file.seek(0)
        
        return Response(
            mem_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment;filename=thermodynamic_properties.xlsx'}
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Failed to generate Excel file.", "details": str(e)}), 500

# app.py dosyanızdaki mevcut export_txt fonksiyonunu bununla değiştirin.

@app.route('/export_txt', methods=['POST'])
def export_txt():
    try:
        points = request.get_json().get('points', [])
        if not points: 
            return jsonify({"error": "No data to export"}), 400
        
        report_str = io.StringIO()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        report_str.write("="*80 + f"\nThermodynamic Properties Report - Generated on {now}\n" + "="*80 + "\n\n")
        
        # Ana tablo
        header = f"{'Point':<15} {'Phase':<25} {'P (kPa)':>15} {'T (°C)':>15} {'h (kJ/kg)':>15} {'s (kJ/kg·K)':>15} {'v (m³/kg)':>15} {'Quality':>15}\n"
        report_str.write(header + "-" * 135 + "\n")
        
        default_text = "Not Available"

        for point in points:
            def fmt(val, precision=4):
                if val is None or (isinstance(val, float) and math.isnan(val)):
                    return default_text
                return f"{val:.{precision}f}"

            line = (f"{point.get('name', default_text):<15} "
                    f"{point.get('phase', default_text):<25} "
                    f"{fmt(point.get('P'), 2):>15} "
                    f"{fmt(point.get('T'), 2):>15} "
                    f"{fmt(point.get('h'), 4):>15} "
                    f"{fmt(point.get('s'), 4):>15} "
                    f"{fmt(point.get('v'), 6):>15} "
                    f"{fmt(point.get('x'), 4):>15}\n")
            report_str.write(line)
        
        # ⭐ SATURATION PROPERTIES TABLOSU
        points_with_sat = [p for p in points if p.get('saturation') and isinstance(p.get('saturation'), dict) and len(p.get('saturation')) > 0]
        
        if points_with_sat:
            report_str.write("\n\n" + "="*160 + "\n")
            report_str.write("SATURATION PROPERTIES\n")
            report_str.write("="*160 + "\n\n")
            
            # Saturation tablo başlıkları
            sat_header = (f"{'Point':<10} "
                         f"{'hf (kJ/kg)':>15} {'hfg (kJ/kg)':>15} {'hg (kJ/kg)':>15} "
                         f"{'sf (kJ/kg·K)':>15} {'sfg (kJ/kg·K)':>15} {'sg (kJ/kg·K)':>15} "
                         f"{'vf (m³/kg)':>15} {'vfg (m³/kg)':>15} {'vg (m³/kg)':>15} "
                         f"{'Tsat/Psat':>20}\n")
            report_str.write(sat_header + "-" * 160 + "\n")
            
            for point in points_with_sat:
                sat = point.get('saturation', {})
                
                def fmt_sat(val, precision=4):
                    if val is None or math.isnan(val):
                        return default_text
                    return f"{val:.{precision}f}"
                
                # Tsat veya Psat bilgisi
                tsat_psat = ""
                if 'T_sat' in sat and not math.isnan(sat.get('T_sat', float('nan'))):
                    tsat_psat = f"Tsat={sat['T_sat']:.2f}°C"
                elif 'P_sat' in sat and not math.isnan(sat.get('P_sat', float('nan'))):
                    tsat_psat = f"Psat={sat['P_sat']:.2f}kPa"
                else:
                    tsat_psat = default_text
                
                sat_line = (f"{point.get('name', default_text):<10} "
                           f"{fmt_sat(sat.get('h_f'), 4):>15} "
                           f"{fmt_sat(sat.get('h_fg'), 4):>15} "
                           f"{fmt_sat(sat.get('h_g'), 4):>15} "
                           f"{fmt_sat(sat.get('s_f'), 4):>15} "
                           f"{fmt_sat(sat.get('s_fg'), 4):>15} "
                           f"{fmt_sat(sat.get('s_g'), 4):>15} "
                           f"{fmt_sat(sat.get('v_f'), 6):>15} "
                           f"{fmt_sat(sat.get('v_fg'), 6):>15} "
                           f"{fmt_sat(sat.get('v_g'), 6):>15} "
                           f"{tsat_psat:>20}\n")
                report_str.write(sat_line)
            
        report_content = report_str.getvalue()
        report_str.close()
        
        return Response(
            report_content,
            mimetype='text/plain',
            headers={'Content-Disposition': 'attachment;filename=properties_report.txt'}
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Failed to generate text file.", "details": str(e)}), 500

@app.route('/calculate_iso_line', methods=['POST'])
def calculate_iso_line():
    data = request.get_json()
    if not data: 
        return jsonify({"error": "Invalid data format"}), 400
    try:
        prop_type, prop_value = data.get('property_type'), float(data.get('property_value'))
        if prop_type != 'P': 
            return jsonify({"error": "Only constant pressure (P) lines are supported."}), 400
        
        p_bar = kpa_to_bar(prop_value)
        tsat = safe_calculate(steamTable.tsat_p, p_bar)
        points_list = []

        if not math.isnan(tsat):
            sL = safe_calculate(steamTable.sL_p, p_bar)
            sV = safe_calculate(steamTable.sV_p, p_bar)
            if not math.isnan(sL): 
                points_list.append({'T': tsat, 's': sL})
            if not math.isnan(sV): 
                points_list.append({'T': tsat, 's': sV})

        temp_range = list(range(1, int(tsat), 5)) + list(range(int(tsat), 801, 10)) if not math.isnan(tsat) else range(1, 801, 10)
        
        for t_c in sorted(list(set(temp_range))):
            s = safe_calculate(steamTable.s_pt, p_bar, t_c)
            if not math.isnan(s):
                points_list.append({'T': t_c, 's': s})

        return jsonify({"points": sorted(points_list, key=lambda p: p['s'])}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": "Failed to calculate iso-line.", "details": str(e)}), 500

@app.route('/calculate_efficiency_custom', methods=['POST'])
def calculate_efficiency_custom():
    data = request.get_json()
    heat_in_processes = data.get('heat_in_processes', [])
    heat_out_processes = data.get('heat_out_processes', [])
    points_map = data.get('points_map', {})

    if not heat_in_processes:
        return jsonify({"error": "At least one heat input process (Q_H) must be defined."}), 400

    total_q_h = 0.0
    total_q_l = 0.0

    try:
        for process in heat_in_processes:
            inlet_point = points_map.get(process['inlet'])
            outlet_point = points_map.get(process['outlet'])
            mass_flow = float(process['mass_flow'])
            if not inlet_point or not outlet_point: raise ValueError(f"Point not found for Q_H process: {process['inlet']} -> {process['outlet']}")
            if 'h' not in inlet_point or 'h' not in outlet_point: raise ValueError(f"Enthalpy (h) is missing for points in process: {process['inlet']} -> {process['outlet']}")
            h_in = inlet_point['h']; h_out = outlet_point['h']; q_h_process = mass_flow * (h_out - h_in)
            if q_h_process > 0: total_q_h += q_h_process
            else: raise ValueError(f"Process {process['inlet']}->{process['outlet']} is defined as heat input, but calculated heat is negative or zero. Check points.")

        for process in heat_out_processes:
            inlet_point = points_map.get(process['inlet'])
            outlet_point = points_map.get(process['outlet'])
            mass_flow = float(process['mass_flow'])
            if not inlet_point or not outlet_point: raise ValueError(f"Point not found for Q_L process: {process['inlet']} -> {process['outlet']}")
            if 'h' not in inlet_point or 'h' not in outlet_point: raise ValueError(f"Enthalpy (h) is missing for points in process: {process['inlet']} -> {process['outlet']}")
            h_in = inlet_point['h']; h_out = outlet_point['h']; q_l_process = mass_flow * (h_out - h_in)
            if q_l_process < 0: total_q_l += abs(q_l_process)
            else: raise ValueError(f"Process {process['inlet']}->{process['outlet']} is defined as heat rejection, but calculated heat is positive or zero. Check points.")

        if total_q_h <= 0:
            return jsonify({"error": "Total heat input (Q_H) must be positive to calculate efficiency."}), 400
        
        efficiency = (1 - total_q_l / total_q_h) * 100

        return jsonify({
            "efficiency": efficiency,
            "total_qh": total_q_h,
            "total_ql": total_q_l
        }), 200

    except (ValueError, TypeError) as e: return jsonify({"error": str(e)}), 400
    except Exception as e: traceback.print_exc(); return jsonify({"error": "An unexpected error occurred during calculation."}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)